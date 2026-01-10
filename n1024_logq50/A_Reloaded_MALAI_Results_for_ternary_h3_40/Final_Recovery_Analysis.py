import openpyxl

# Define the hamming weights
hamming_weights = list(range(20, 41))

# Function to read and process a recovery file
def process_file(filename):
    try:
        with open(filename, 'r') as file:
            lines = file.readlines()
    except FileNotFoundError:
        print(f"File not found: {filename}")
        return None, None
    
    min_no_mod = {}
    min_samples = {}

    current_hamming_weight = None
    for line in lines:
        line = line.strip()
        if line.startswith("************** Hamming weight ="):
            current_hamming_weight = int(line.split('=')[1].strip().split()[0])
        elif line.startswith("NoMod"):
            no_mod_percentage, samples_info = line.split('%  ==>')
            no_mod_percentage = int(no_mod_percentage.replace('NoMod', '').strip())
            samples_info = samples_info.strip()
            if samples_info != "Not achieved":
                try:
                    samples = int(samples_info.split()[0])
                    if current_hamming_weight not in min_no_mod or min_no_mod[current_hamming_weight] > no_mod_percentage:
                        min_no_mod[current_hamming_weight] = no_mod_percentage
                        min_samples[current_hamming_weight] = samples
                except ValueError:
                    pass

    return min_no_mod, min_samples

# Process the files for 80%, 90%, and 100% recovery
min_no_mod_10, min_samples_10 = process_file("Minimum_samples_required_for_10_Percent_Recovery_for_h3_40_ternary.txt")
min_no_mod_100, min_samples_100 = process_file("Minimum_samples_required_for_100_Percent_Recovery_for_h3_40_ternary.txt")

# Create a new Excel workbook and sheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Recovery Analysis"

# Write the header row
ws.append([
    "Hamming Weight",
    "For 10% Recovery Minimum NoMod%", "For 10% Recovery Minimum #Samples Required",
    "For 100% Recovery Minimum NoMod%", "For 100% Recovery Minimum #Samples Required"
])

# Write the data rows
for hw in hamming_weights:
    row = [hw]

    # 10% recovery
    if hw in min_no_mod_10 and min_no_mod_10[hw] is not None and min_samples_10[hw] is not None:
        row.append(min_no_mod_10[hw])
        row.append(min_samples_10[hw])
    else:
        row.extend([-1, -1])


    # 100% recovery
    if hw in min_no_mod_100 and min_no_mod_100[hw] is not None and min_samples_100[hw] is not None:
        row.append(min_no_mod_100[hw])
        row.append(min_samples_100[hw])
    else:
        row.extend([-1, -1])

    ws.append(row)

# Save the initial workbook
output_filename = "Final_Recovery_Analysis_N_1024_Q_50_ternary.xlsx"
wb.save(output_filename)
print(f"Initial results have been written to {output_filename}")

# Load the workbook again to perform transposition
wb = openpyxl.load_workbook(output_filename)
ws = wb.active

# Create a new workbook for the transposed data
output_wb = openpyxl.Workbook()
output_ws = output_wb.active
output_ws.title = "Transposed Recovery Analysis"

# Transpose the data
for i, row in enumerate(ws.iter_rows(values_only=True)):
    for j, value in enumerate(row):
        output_ws.cell(row=j+1, column=i+1, value=value)

# Save the transposed workbook
transposed_filename = "Final_Transposed_Recovery_Analysis_N_1024_Q_50_ternary.xlsx"
output_wb.save(transposed_filename)

print(f"Transposed data has been written to {transposed_filename}")